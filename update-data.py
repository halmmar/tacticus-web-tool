#!/usr/bin/env python3

import numpy as np
import pandas as pd
import xlsxwriter
import subprocess
import openpyxl
import json
import copy
from collections import OrderedDict
from io import StringIO

file = 'TACTICUS-0.5.4.xlsx'

# Map between Towen and SP names
nameMap = {
    "Abaddon the despolier": "Abaddon",
    "Aun'shi": "Aun'Shi",
    "Aleph-null": "Aleph-Null",
    "Ancient Thoread": "Thoread",
    "Boss Gulgortz": "Gulgortz",
    "Brother Burchard": "Burchard",
    "Brother Jaeger": "Jaeger",
    "Castellan creed": "Creed",
    "Command-Link drone": "Command-Link Drone",
    "Commissar Yarrick": "Yarrick",
    "Exitor-Rho-1.15/x": "Exitor-Rho",
    "Haarken Worldclaimer": "Haarken",
    "High Marshal Helbrecht": "Helbrecht",
    "Kut skoden": "Kut",
    "Marneus Calgar": "Calgar",
    "Nauseous rotbone": "Rotbone", # Also modified in the Javascript since SP named him "Nauseus"
    "Parasite Of Mortrex": "Parasite of Mortrex",
    "Re'Vas": "Re'vas",
    "ShadowSun": "Shadowsun",
    "Sho'Syl": "Sho'syl",
    "Sword Brother Godswyl": "Godswyl",
    "Tan Gi’da": "Tan Gi'da",
    "Thaddeus noble": "Thaddeus",
    "Varro Tigurius": "Tigurius",
    "Necron warrior": "Necron Warrior",
    "Screamer of Tzeentch": "Screamer",
    "Hormagunt": "Hormagaunt",
    "Skitarii Vanguards": "Skitarii Vanguard",
    "Cadian Guardsman": "Cadian"
}

def data_frame_from_xlsx(xlsx_file, range_name, headerColumn=None):
    """ Get a single rectangular region from the specified file.
    range_name can be a standard Excel reference ('Sheet1!A2:B7') or
    refer to a named region ('my_cells')."""
    wb = openpyxl.load_workbook(xlsx_file, data_only=True, read_only=True)
    columns = None
    if '!' in range_name:
        # passed a worksheet!cell reference
        ws_name, reg = range_name.split('!')
        if ws_name.startswith("'") and ws_name.endswith("'"):
            # optionally strip single quotes around sheet name
            ws_name = ws_name[1:-1]
        ws = wb[ws_name]
    else:
        # passed a named range; find the cells in the workbook
        full_range = wb.defined_names[range_name]
        if full_range is None:
            raise ValueError(
                'Range "{}" not found in workbook "{}".'.format(range_name, xlsx_file)
            )
        print(full_range.value)
        print(type(full_range))
        # convert to list (openpyxl 2.3 returns a list but 2.4+ returns a generator)
        destinations = list(full_range.destinations)
        print(destinations)
        if len(destinations) > 1:
            raise ValueError(
                'Range "{}" in workbook "{}" contains more than one region.'
                .format(range_name, xlsx_file)
            )
        ws, reg = destinations[0]
        # convert to worksheet object (openpyxl 2.3 returns a worksheet object 
        # but 2.4+ returns the name of a worksheet)
        if isinstance(ws, str):
            ws = wb[ws]
    region = ws[reg]
    if headerColumn is None:
        pass
    else:
        heads = reg.split("$")
        if type(headerColumn) is int:
            headRow = headerColumn
        elif type(headerColumn) is bool:
            headRow = int(heads[2][:-1])-1
        else:
            raise Exception("...")
        headRange = "$%s$%d:$%s$%d" % (heads[1],headRow,heads[3],headRow)
        headRegion = ws[headRange]
        columns = [(cell.value.strip() if isinstance(cell.value,str) else cell.value) for cell in headRegion[0]]
    # an anonymous user suggested this to catch a single-cell range (untested):
    # if not isinstance(region, 'tuple'): df = pd.DataFrame(region.value)
    df = pd.DataFrame(([cell.value for cell in row] for row in region), columns=columns)
    return df

bossSummons = ["Stormboy"]

tb_Characters = data_frame_from_xlsx(file, 'tb_Characters', 1)
tb_Pierce = data_frame_from_xlsx(file, 'UL_Tables!$Z$2:$AA$24', True)
tb_Gear = data_frame_from_xlsx(file, 'UL_Tables!$K$4:$O$21', True)

# Updated for new factions; also update the faction map below
tb_Equipment = data_frame_from_xlsx(file, 'Equipment!$B$3:$AX$151', 2)
# Updated for new characters
tb_CharacterAbilities = data_frame_from_xlsx(file, 'wb_abilities!$A$4:$U$89', True) # Passives

tb_Abilities = data_frame_from_xlsx(file, 'UL_Tables!$AI$3:$AM$52', True) # Not updated for new characters
tb_Abilities_1_50 = [level for (level,factor,archi,azkor,titus) in tb_Abilities.itertuples(index=False)]
# tb_SummonsList = data_frame_from_xlsx(file, 'SummonData!$A$2:$A$30', False)

for i in range(1,51):
    if tb_Abilities_1_50[i-1] != i:
        raise Exception("Did not find the correct table. i=%d, tb_Abilities=%d" % ( i, tb_Abilities_1_50[i-1]))
tb_Abilities_factor = [factor for (level,factor,archi,azkor,titus) in tb_Abilities.itertuples(index=False)]
tb_Abilities_factor_archimatos = [archi for (level,factor,archi,azkor,titus) in tb_Abilities.itertuples(index=False)]
tb_Abilities_factor_azkor = [azkor for (level,factor,archi,azkor,titus) in tb_Abilities.itertuples(index=False)]
tb_Abilities_factor_titus = [titus for (level,factor,archi,azkor,titus) in tb_Abilities.itertuples(index=False)]
summonsNames = set()
for index, row in []: # tb_SummonsList.iterrows():
    if row[0] and not row[0] in bossSummons:
        if row[0] == "Command-Link drone":
            row[0] = "Command-Link Drone"
        summonsNames.add(row[0])
# Legendary Events
tb_LegendaryEvent = data_frame_from_xlsx(file, 'YourUnits!$A$1:$EA$99')
# 76=Excel column BY, which conflicts with the query language used by Towen. The column is empty and skipped...
tb_LegendaryEvent.drop(columns=[76], axis=1, inplace=True)
tb_LegendaryEvent.to_excel("tmp.xlsx", header=False, index=False)
tb_LegendaryEvent = pd.read_excel("tmp.xlsx", header=None)

lastIndexCharacter = 5
while isinstance(tb_LegendaryEvent[0][lastIndexCharacter], str):
    lastIndexCharacter += 1
# Passives
characterAbilities = dict([(tpl[0],tpl[1:]) for tpl in tb_CharacterAbilities.itertuples(index=False)])

indexLegendaryEvent=17
legendaryEvents = {}
legendaryEventCharacters = {}
latestLegendaryEvent = None
while indexLegendaryEvent in tb_LegendaryEvent and tb_LegendaryEvent[indexLegendaryEvent][0]:
    legendaryEventName = tb_LegendaryEvent[indexLegendaryEvent][0].split("-")[1].strip()
    latestLegendaryEvent = legendaryEventName
    items = tb_LegendaryEvent.iloc[3:5,indexLegendaryEvent:indexLegendaryEvent+18]
    points = list(tb_LegendaryEvent.iloc[2,indexLegendaryEvent:indexLegendaryEvent+18])
    items.replace("No", "No ",inplace=True)
    items.replace("Eligiable", "Eligible",inplace=True)
    items.replace("Eligible", "!", inplace=True)
    items.fillna("",inplace=True)
    itemsList = list(items.sum())
    alphaNames = itemsList[0:6]
    betaNames = itemsList[6:12]
    gammaNames = itemsList[12:18]
    pointsList = list(points)
    
    alpha = dict(zip(alphaNames,[[p,0] for p in pointsList[0:6]]))
    beta = dict(zip(betaNames,[[p,0] for p in pointsList[6:12]]))
    gamma = dict(zip(gammaNames,[[p,0] for p in pointsList[12:18]]))
    for index, row in tb_LegendaryEvent.iterrows():
        if index < 5:
            continue
        if not row[0]:
            continue
        if not isinstance(row[0],str):
            continue
        row[0] = nameMap.get(row[0], row[0])
        if row[0] not in legendaryEventCharacters:
            legendaryEventCharacters[row[0]] = {}
        characterAlpha = []
        characterBeta = []
        characterGamma = []
        for i in range(0,6):
            if row[indexLegendaryEvent+i]:
                characterAlpha += [alphaNames[i]]
            if row[indexLegendaryEvent+6+i]:
                characterBeta += [betaNames[i]]
            if row[indexLegendaryEvent+12+i]:
                characterGamma += [gammaNames[i]]
        for (missionList,missionPoints) in [(characterAlpha,alpha),(characterBeta,beta),(characterGamma,gamma)]:
            if "!" not in missionList:
                missionList.clear()
            for mission in missionList:
                missionPoints[mission][1] += 1
        if legendaryEventName in legendaryEventCharacters[row[0]]:
            raise Exception(row[0], "appears twice in the legendary events (YourUnits)")
        legendaryEventCharacters[row[0]][legendaryEventName] = {"alpha": characterAlpha, "beta": characterBeta, "gamma": characterGamma}
                
    legendaryEvents[legendaryEventName] = {"alpha": alpha, "beta": beta, "gamma": gamma}
    indexLegendaryEvent += 19

# Boss stats
bossStats = {
  # Stats are an array [0]=L1, etc. The first element is the stat and the second, the season which the data was collected from
  "Avatar": {
      "armour": [None,1029,1288,1697,2231],
      "checkbox": 0.7,
      "traits": ["big","daemon"]
  },
  "Magnus": {
      "armour": [None,None,989,1303,1713],
      "checkbox": 1.0,
      "traits": ["big","daemon","flying","psyker"]
  },
  "Cawl": {
      "armour": [None,880,1102,None,None],
      "checkbox": 0.7,
      "traits": ["big","mech"]
  },
  "Ghazghkull": {
      "armour": [352,465,582,660,1008],
      "checkbox": 0.5,
      "traits": ["big","mech"]
  },
  "Mortarion": {
      "armour": [2337,4385,None,5095,6698],
      "checkbox": 0.1,
      "traits": ["big","daemon","psyker","revoltingly resilient","punishes melee"]
  },
  "Tervigon": { # Gorgon
      "armour": [666, 1020, 1559, 1559, 1559],
      "checkbox": 0.4,
      "traits": ["big","psyker","flying"]
  },
  "Hive Tyrant": { # Gorgon
      "armour": [666, 720, 1559, 1559, 1559],
      "checkbox": 0.4,
      "traits": ["big","psyker","flying"]
  },
  "Szarekh": {
      "armour": [367, 485, None, 800, 1052],
      "checkbox": 0.35,
      "traits": ["big","mech"]
  },
  "Screamer-Killer": {
      "armour": [830, 1323, None, 1809, 2379],
      "checkbox": 0.85,
      "traits": ["big"]
  },
  "Rogal Dorn": {
      "armour": [998,1319,2861,2861,2861],
      "checkbox": 0.7,
      "traits": ["big","mech","ablative plating","weaker rear"]
  }
}

pierce = dict([(k.lower(),v) for (k,v) in tb_Pierce.itertuples(index=False)])
gear = [(rank,int(gearLevel)) for (rarity,rank,rankMetal,rankLevel,gearLevel) in tb_Gear.itertuples(index=False)]

factionMap = {
    "Adepta Sororitas": "Adepta Sororitas",
    "ADEPTA SORORITAS": "Adepta Sororitas",
    "Adeptus Mechanicus": "Adeptus Mechanicus",
    "ADEPTUS MECHANICUS": "Adeptus Mechanicus",
    "ASTRA MILITARUM": "Astra militarum",
    "Astra Miltarum": "Astra militarum",
    "Astra Militarum": "Astra militarum",
    "BLACK LEGION": "Black Legion",
    "Black Legion": "Black Legion",
    "BLACK TEMPLARS": "Black templars",
    "Black Templars": "Black templars",
    "BLOOD ANGELS": "Blood Angels",
    "Blood Angels": "Blood Angels",
    "DARK ANGELS": "Dark Angels",
    "Dark Angels": "Dark Angels",
    "Death guard": "Death Guard",
    "DEATH GUARD": "Death Guard",
    "Genestealer Cults": "Genestealer Cults",
    "GENESTEALER CULTS": "Genestealer Cults",
    "ORKS": "Orks",
    "Orks": "Orks",
    "Ork": "Orks",
    "NECRONS": "Necrons",
    "Necrons": "Necrons",
    "AELDARI": "Aeldari",
    "Aeldari": "Aeldari",
    "T'AU": "T'au Empire",
    "T'AU EMPIRE": "T'au Empire",
    "T'au Empire": "T'au Empire",
    "Space Wolves": "Space Wolves",
    "SPACE WOLVES": "Space Wolves",
    "THOUSAND SONS": "Thousand Sons",
    "Thousand Sons": "Thousand Sons",
    "Tyranids": "Tyranids",
    "TYRANIDS": "Tyranids",
    "Ultramarines": "Ultramarines",
    "ULTRAMARINES": "Ultramarines",
    "WORLD EATERS": "World Eaters"
}
allianceMap = {
    "XENOS": "Xenos",
    "Xenos": "Xenos",
    "Xenons": "Xenos",
    "CHAOS": "Chaos",
    "Chaos": "Chaos",
    "IMPERIAL": "Imperial",
    "Imperial": "Imperial",
    None: "",
    0: ""
}

eqDict = {}
for index, eq in tb_Equipment.iterrows():
    if eq["Item Type"] is None:
        continue
    itemTypeStr = eq["Item Type"].lower().replace(" ","_")
    if itemTypeStr not in eqDict:
        eqDict[itemTypeStr] = {}
    itemType = eqDict[itemTypeStr]
    factions = []
    for faction in factionMap.keys():
        if faction in eq and eq[faction]:
            factions += [factionMap[faction]]
    itemInternalType = eq[0:1].iloc(0) # Gun, Knife, Both
    chance = eq["Chance"]
    if np.isnan(chance):
        chance = int("Both"==itemInternalType)
    else:
        chance = int(chance)
    if "booster" in itemTypeStr:
        chance = 1
    if chance in itemType:
        itemType[chance]["factions"] = list(OrderedDict.fromkeys(itemType[chance]["factions"]+factions))
    else:
        itemType[chance] = {
            "factions": factions
        }
    itemChance = itemType[chance]
    rarityStr = eq["Rarity"].lower()
    if rarityStr not in itemChance:
        stat1 = [int(x) for x in eq[26:37] if not np.isnan(x)]
        stat2 = [int(x) for x in eq[38:49] if not np.isnan(x)]
        if itemTypeStr=="defense":
            itemChance[rarityStr] = {
                "health": stat1,
                "armour": stat2
            }
        else:
            itemChance[rarityStr] = {
                {
                    "crit": "crit-dmg",
                    "crit_booster": "crit-dmg",
                    "block": "block",
                    "block_booster": "block"
                }[itemTypeStr]: stat1
            }

summonMap = {
    "Inceptor": {"characters": ["Bellator"], "ability": "active"},
    "Bloodletter": {"characters": ["Archimatos"], "ability": "both"},
    "Grot": {"characters": ["Snotflogga"], "ability": "both"},
    "Grot Tank": {"characters": ["Gibbascrapz"], "ability": "active"},
    "Scarab Swarm": {"characters": ["Aleph-Null"], "ability": "active"},
    "Necron Warrior": {"characters": ["Aleph-Null"], "ability": "active"},
    "Cadian": {"characters": ["Yarrick","Creed"], "ability": "active"},
    "Colour Sergeant Kell": {"characters": ["Creed"], "ability": "passive"},
    "Geminae Superia": {"characters": ["Celestine"], "ability": "passive"},
    "Pox Walkers": {"characters": ["Corrodius"], "ability": "active"},
    "Shield Drone": {"characters": ["Re'vas"], "ability": "active"},
    "MV71 Sniper Drone": {"characters": ["Sho'syl"], "ability": "passive"},
    "Command-Link Drone": {"characters": ["Shadowsun"], "ability": "active", "stats": [30,14,10]},
    "Pink Horror": {"characters": ["Abraxas"], "ability": "active", "stats": [36,9,0]},
    "Blue Horror": {"characters": ["Abraxas"], "ability": "active", "stats": [21,18,0]},
    "Screamer": {"characters": ["Abraxas"], "ability": "active", "stats": [45,18,0]},
    "Tyranid Warrior": {"characters": ["Winged Prime"], "ability": "active"},
    "Hormagaunt": {"characters": ["Winged Prime"], "ability": "passive"},
    "Termagant": {"characters": []}, # Why is this in Towen's sheet?
    "Ripper Swarm": {"characters": ["Parasite of Mortrex"], "ability": "both"},
    "Skitarii Vanguard": {"characters": ["Tan Gi'da"], "ability": "passive"},
    "Jump Pack Intercessor": {"characters": ["Mataneo"], "ability": "passive"},
}

def toJSON(rows):
    characters = {}
    bosses = {}
    summons = {}
    bossNames = ["TERVIGON", "Tervigon"]
    foundMobs = False
    for index, row in rows.iterrows():
        if row.iloc(0) in ["Mobs","GRBosses"]:
            foundMobs = True
        if row.iloc(0) == "Summons":
            foundMobs = False
        if foundMobs:
            continue
        name = nameMap.get(row.Name, row.Name)
        if row.Faction in bossNames:
            continue
        if row.Alliance in bossNames:
            break
        if row.Name is None or np.isnan(row["Melee Hits"]) or np.isnan(row.Health) or row["Melee Damage"]==0:
            continue
        eqCount = {"crit": 0, "defense": 0, "crit_booster": 0, "block": 0, "block_booster": 0}
        for eqNum in ["Equipement slot 1", "Equipement slot 2", "Equipemnt slot 3"]:
            if row[eqNum] is None or type(row[eqNum]) is not str:
                continue
            eq = row[eqNum].lower().replace(" ", "_")
            if eq in eqCount:
                eqCount[eq] += 1
        traits = []
        for trait in ["Trait %d" % n for n in range(1,6)]:
            if row[trait] is None or type(row[trait]) is not str:
                continue
            traitStripped = row[trait].lower().strip().replace("living metsl", "living metal")
            if traitStripped in traits:
                continue
            traits += [traitStripped]
        if row.Name in ["Shield Drone","Termagant"] and "summon" not in traits:
            traits += ["summon"]
        toDelete = []
        for key in eqCount.keys():
            if not eqCount[key]:
                toDelete += [key]
        for key in toDelete:
            del eqCount[key]
        passiveData = []
        if row.Name in characterAbilities:
            passiveData = [x for x in characterAbilities[row.Name][17:20] if not (x is None or np.isnan(x))]
        activeData = []
        if row.Name in characterAbilities:
            activeData = [x for x in characterAbilities[row.Name][12:17] if not (x is None or np.isnan(x))]
        data = {"faction": factionMap[row.Faction], "alliance": allianceMap[row.Alliance], "health": row.Health, "damage": row.Damage, "traits": traits, "armour": row.Armour, "melee": {"pierce": pierce[row["Melee Damage"].lower()], "hits": int(row["Melee Hits"]), "type": row["Melee Damage"].lower()}}
        if row["Ranged Hits"] > 0:
            data["ranged"] = {"pierce": pierce[row["Ranged Damage"].lower()], "hits": int(row["Ranged Hits"]), "type": row["Ranged Damage"].lower()}
        if row["Initial rarity"] in ["Common", "Uncommon", "Rare", "Epic", "Legendary"]:
            data["passive"] = passiveData
            data["active"] = activeData
            data["equipment"] = eqCount
            # data["legendary-event"] = legendaryEventCharacters[name]
            characters[name] = data
        #elif row.Name[0:2] in ["L1", "L2", "L3", "L4"]:
        #    bosses[row.Name] = data
        elif "summon" in traits:
            if name not in summonMap:
                raise Exception(name + " " + str(summonMap.keys()) + "\n" + str(row))
            for character in summonMap[name]["characters"]:
                abilityUsed = summonMap[name]["ability"]
                for ability in ["active","passive"]:
                    if not (abilityUsed=="both" or abilityUsed==ability):
                        continue
                    stats = summonMap[name].get("stats",[])
                    if len(stats)==0:
                        stats = characters[character][ability]
                    data["health"] = stats[0]
                    data["damage"] = stats[1]
                    data["armour"] = stats[2] if len(stats)>=3 else 0
                    data["summoner"] = character
                    data["summonerAbility"] = ability
                    summonName = name
                    if len(summonMap[name]["characters"]) > 1:
                        summonName += " ("+character+")"
                    if abilityUsed == "both":
                        summonName += " " + ability[0].upper()
                    summons[summonName] = data.copy()
        else:
            continue
    for name in summonsNames:
        if name not in summons:
            raise Exception("%s is not in the list of summons" % name)
    return {"characters": characters, "bosses": bossStats, "summons": summons, "gear": gear, "abilities_factor": tb_Abilities_factor, "archimatos_ability_factor": tb_Abilities_factor_archimatos, "azkor_ability_factor": tb_Abilities_factor_azkor, "titus_ability_factor": tb_Abilities_factor_titus, "equipment": eqDict, "version": file.split("-")[1].strip().replace(".xlsx", "")} # , "legendary-events": legendaryEvents, "latest-legendary-event": latestLegendaryEvent,
with open("tacticus.json", "w") as fout:
    fout.write(json.dumps(toJSON(tb_Characters), sort_keys=True, indent=2))
